from flask import Flask, render_template, request, send_file, redirect, url_for, abort
from logic.billtracker import build_excel
from logic.toc_linker import process_pdf as process_toc
from logic.ocr_checker import analyze_pdf_bytes
import os
import io, zipfile
from flask import jsonify
from logic.billtracker import build_excel, build_excel_from_rows

# Excel writing
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = (
    200 * 1024 * 1024
)  # 200 MB — bumped for batch PDF uploads

# Optional very simple shared password
APP_PASSWORD = os.getenv("APP_PASSWORD", "")


def require_auth(fn):
    def wrapper(*args, **kwargs):
        if APP_PASSWORD:
            if request.method == "POST":
                if request.form.get("password") != APP_PASSWORD:
                    abort(401)
            else:
                pass
        return fn(*args, **kwargs)

    wrapper.__name__ = fn.__name__
    return wrapper


# ── Billing Tracker ───────────────────────────────────────────────────────────

@app.post("/billtracker-json")
def bill_post_json():
    data = request.get_json(silent=True) or {}
    rows = data.get("rows", [])
    if not rows:
        return jsonify({"error": "No rows provided"}), 400
    out = build_excel_from_rows(rows)
    return send_file(out, as_attachment=True, download_name="pdf_page_counts.xlsx")


@app.get("/")
def index():
    return render_template("index.html", require_pw=bool(APP_PASSWORD))


@app.get("/billtracker")
def bill_get():
    return render_template("billtracker.html", require_pw=bool(APP_PASSWORD))


@app.post("/billtracker")
@require_auth
def bill_post():
    files = request.files.getlist("pdfs")
    if not files:
        return redirect(url_for("bill_get"))
    out = build_excel(files)
    return send_file(out, as_attachment=True, download_name="pdf_page_counts.xlsx")


# ── TOC Linker ────────────────────────────────────────────────────────────────

@app.get("/toc-linker")
def toc_get():
    return render_template("toc_linker.html", require_pw=bool(APP_PASSWORD))


@app.post("/toc-linker")
@require_auth
def toc_post():
    pdfs = request.files.getlist("pdfs")
    rng  = request.form.get("range", "").strip()

    if not pdfs or not rng:
        return redirect(url_for("toc_get"))

    if len(pdfs) == 1:
        pdf  = pdfs[0]
        out  = process_toc(pdf, rng)
        name = (pdf.filename or "output.pdf").rsplit(".", 1)[0] + "_TOCLinked.pdf"
        return send_file(out, as_attachment=True, download_name=name)

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for pdf in pdfs:
            if not pdf or not pdf.filename:
                continue
            try:
                out      = process_toc(pdf, rng)
                out_name = pdf.filename.rsplit(".", 1)[0] + "_TOCLinked.pdf"
                zf.writestr(out_name, out.getvalue())
            except Exception as e:
                err_name = pdf.filename.rsplit(".", 1)[0] + "_ERROR.txt"
                zf.writestr(err_name, f"Failed to process: {pdf.filename}\n{e}")

    zip_buf.seek(0)
    return send_file(
        zip_buf,
        as_attachment=True,
        download_name="TOCLinked_Batch.zip",
        mimetype="application/zip",
    )


# ── OCR Checker ───────────────────────────────────────────────────────────────

@app.get("/ocr-checker")
def ocr_get():
    return render_template("ocr_checker.html", require_pw=bool(APP_PASSWORD))


@app.post("/ocr-checker")
def ocr_post():
    files = request.files.getlist("pdfs")
    if not files:
        return jsonify({"error": "No files provided"}), 400

    results = []
    for f in files:
        if not f or not f.filename:
            continue
        file_bytes = f.read()
        result = analyze_pdf_bytes(file_bytes, f.filename)
        results.append(result)

    out = _build_ocr_excel(results)
    return send_file(out, as_attachment=True, download_name="ocr_report.xlsx")


def _build_ocr_excel(results: list) -> io.BytesIO:
    """Build a nicely formatted Excel report from OCR analysis results."""
    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"

    # Colour palette
    HDR_FILL   = PatternFill("solid", fgColor="2F5496")
    RED_FILL   = PatternFill("solid", fgColor="F4CCCC")
    AMBER_FILL = PatternFill("solid", fgColor="FCE5CD")
    GREEN_FILL = PatternFill("solid", fgColor="D9EAD3")
    ALT_FILL   = PatternFill("solid", fgColor="F3F6FB")
    WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")

    HDR_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    BOLD_FONT = Font(bold=True, name="Calibri", size=11)
    STD_FONT  = Font(name="Calibri", size=11)

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")

    summary_headers = [
        "Filename", "Total Pages",
        "Needs OCR Pages", "Needs OCR Total",
        "Review Pages", "Review Total",
        "OK", "Overall Status", "Error"
    ]
    col_widths_sum = [45, 13, 50, 15, 50, 13, 10, 18, 30]

    ws_sum.row_dimensions[1].height = 22
    for col, (hdr, w) in enumerate(zip(summary_headers, col_widths_sum), start=1):
        cell = ws_sum.cell(row=1, column=col, value=hdr)
        cell.fill      = HDR_FILL
        cell.font      = HDR_FONT
        cell.alignment = center
        cell.border    = border
        ws_sum.column_dimensions[cell.column_letter].width = w

    for r_idx, res in enumerate(results, start=2):
        is_alt = (r_idx % 2 == 0)
        base_fill = ALT_FILL if is_alt else WHITE_FILL

        if res.get("error"):
            overall = "ERROR"
        elif res["needs_ocr"] > 0:
            overall = "NEEDS OCR"
        elif res["review"] > 0:
            overall = "REVIEW"
        else:
            overall = "OK"

        needs_ocr_pages = ", ".join(str(p["page"]) for p in res["pages"] if p["status"] == "NEEDS OCR")
        review_pages    = ", ".join(str(p["page"]) for p in res["pages"] if p["status"] == "REVIEW")

        row_data = [
            res["filename"],
            res["total"],
            needs_ocr_pages,
            res["needs_ocr"],
            review_pages,
            res["review"],
            res["ok"],
            overall,
            res.get("error") or "",
        ]

        for c_idx, value in enumerate(row_data, start=1):
            cell = ws_sum.cell(row=r_idx, column=c_idx, value=value)
            cell.font      = STD_FONT
            cell.border    = border
            cell.alignment = center if c_idx not in (1, 3, 5) else left
            if c_idx in (3, 5):
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

            # Overall Status column (col 8)
            if c_idx == 8:
                if overall == "NEEDS OCR":
                    cell.fill = RED_FILL
                    cell.font = Font(bold=True, name="Calibri", size=11, color="990000")
                elif overall == "REVIEW":
                    cell.fill = AMBER_FILL
                    cell.font = Font(bold=True, name="Calibri", size=11, color="7F4F00")
                elif overall == "OK":
                    cell.fill = GREEN_FILL
                    cell.font = Font(bold=True, name="Calibri", size=11, color="1C4D1C")
                else:
                    cell.fill = base_fill
            # Needs OCR pages (col 3) and total (col 4)
            elif c_idx in (3, 4) and res["needs_ocr"] > 0:
                cell.fill = RED_FILL
            # Review pages (col 5) and total (col 6)
            elif c_idx in (5, 6) and res["review"] > 0:
                cell.fill = AMBER_FILL
            else:
                cell.fill = base_fill

        # Auto-height for wrapped page-list cells
        ws_sum.row_dimensions[r_idx].height = 30

    ws_sum.freeze_panes = "A2"

    # ── Sheet 2: Page Detail ──────────────────────────────────────────────────
    ws_det = wb.create_sheet("Page Detail")

    detail_headers = [
        "Filename", "Page #", "Status",
        "Char Count", "Image Count", "Image Contains Text"
    ]
    col_widths_det = [45, 9, 12, 13, 14, 20]

    ws_det.row_dimensions[1].height = 22
    for col, (hdr, w) in enumerate(zip(detail_headers, col_widths_det), start=1):
        cell = ws_det.cell(row=1, column=col, value=hdr)
        cell.fill      = HDR_FILL
        cell.font      = HDR_FONT
        cell.alignment = center
        cell.border    = border
        ws_det.column_dimensions[cell.column_letter].width = w

    det_row = 2
    for res in results:
        if res.get("error") or not res["pages"]:
            continue
        for page in res["pages"]:
            is_alt   = (det_row % 2 == 0)
            base_fill = ALT_FILL if is_alt else WHITE_FILL
            status   = page["status"]

            row_data = [
                res["filename"],
                page["page"],
                status,
                page["char_count"],
                page["image_count"],
                "Yes" if page.get("image_has_text") else "No",
            ]
            for c_idx, value in enumerate(row_data, start=1):
                cell = ws_det.cell(row=det_row, column=c_idx, value=value)
                cell.font      = STD_FONT
                cell.border    = border
                cell.alignment = center if c_idx != 1 else left

                if c_idx == 3:
                    if status == "NEEDS OCR":
                        cell.fill = RED_FILL
                        cell.font = Font(bold=True, name="Calibri", size=11, color="990000")
                    elif status == "REVIEW":
                        cell.fill = AMBER_FILL
                        cell.font = Font(bold=True, name="Calibri", size=11, color="7F4F00")
                    elif status == "OK":
                        cell.fill = GREEN_FILL
                        cell.font = Font(bold=True, name="Calibri", size=11, color="1C4D1C")
                    else:
                        cell.fill = base_fill
                else:
                    cell.fill = base_fill

            det_row += 1

    ws_det.freeze_panes = "A2"
    ws_det.auto_filter.ref = f"A1:F{det_row - 1}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=10000, debug=True)
