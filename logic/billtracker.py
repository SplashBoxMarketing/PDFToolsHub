# logic/billtracker.py
import io, os
from datetime import date
import pandas as pd

# from pypdf import PdfReader  # <- no longer needed for counting
from openpyxl import load_workbook
import fitz  # PyMuPDF


def build_excel_from_rows(rows):
    # rows = [{"Filename": "...", "Page Count": 12, "Complete Y/N": "", "Note": ""}, ...]
    today = date.today().strftime("%Y-%m-%d")
    norm = []
    for r in rows:
        norm.append(
            {
                "Filename": r.get("Filename") or r.get("filename") or "file.pdf",
                "Page Count": r.get("Page Count") or r.get("page_count") or "",
                "Complete Y/N": r.get("Complete Y/N", ""),
                "Note": r.get("Note", ""),
                "Date Received": r.get("Date Received") or today,
                "Date Complete": r.get("Date Complete", ""),
            }
        )

    df = pd.DataFrame(
        norm,
        columns=[
            "Filename",
            "Page Count",
            "Complete Y/N",
            "Note",
            "Date Received",
            "Date Complete",
        ],
    )

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False)
    buf.seek(0)
    wb = load_workbook(buf)
    ws = wb.active
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_len + 2
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def _page_count_from_filestorage(fs) -> int | None:
    try:
        # Read the uploaded file fully into memory once
        data = fs.read()
        if not data:
            return None
        # Use PyMuPDF: very memory-efficient page count
        doc = fitz.open(stream=data, filetype="pdf")
        pc = doc.page_count
        doc.close()
        return pc
    except Exception:
        return None
    finally:
        # Reset stream in case anything upstream reuses it (defensive)
        try:
            fs.stream.seek(0)
        except Exception:
            pass


def build_excel(file_storages, filename_col="Filename"):
    rows = []
    today = date.today().strftime("%Y-%m-%d")
    for fs in file_storages:
        name = os.path.basename(fs.filename or "file.pdf")
        pc = _page_count_from_filestorage(fs)
        rows.append(
            {
                filename_col: name,
                "Page Count": pc if isinstance(pc, int) else "ERROR",
                "Complete Y/N": "",
                "Note": "",
                "Date Received": today,
                "Date Complete": "",
            }
        )

    df = pd.DataFrame(
        rows,
        columns=[
            filename_col,
            "Page Count",
            "Complete Y/N",
            "Note",
            "Date Received",
            "Date Complete",
        ],
    )

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False)

    # Auto-fit columns
    buf.seek(0)
    wb = load_workbook(buf)
    ws = wb.active
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_len + 2
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out
